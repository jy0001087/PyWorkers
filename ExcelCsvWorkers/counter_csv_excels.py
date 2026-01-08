import os
import argparse
from collections import defaultdict

# 用途：
# 统计指定路径下（单个文件或目录）所有 CSV / Excel 文件的行数
# 如果是目录，则按目录分组显示统计结果

CSV_EXTS = {".csv"}
EXCEL_EXTS = {".xls", ".xlsx", ".xlsm", ".xlsb"}


def count_csv_lines(file_path):
    """使用二进制读取，高效统计 CSV 行数"""
    count = 0
    with open(file_path, "rb") as f:
        buf_size = 1024 * 1024
        read_f = f.raw.read
        buf = read_f(buf_size)
        while buf:
            count += buf.count(b"\n")
            buf = read_f(buf_size)
    return count


def count_excel_lines(file_path):
    """统计 Excel 物理行数"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext in {".xlsx", ".xlsm", ".xls"}:
        from openpyxl import load_workbook
        # 注意：.xls 建议使用 xlrd，这里沿用原逻辑
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        return ws.max_row
    elif ext == ".xlsb":
        from pyxlsb import open_workbook
        with open_workbook(file_path) as wb:
            with wb.get_sheet(1) as sheet:
                return sum(1 for _ in sheet.rows())
    return 0


def count_lines(file_path):
    """根据文件类型自动选择统计方式"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext in CSV_EXTS:
        return count_csv_lines(file_path)
    elif ext in EXCEL_EXTS:
        return count_excel_lines(file_path)
    return 0


def process_single_file(file_path):
    """处理单个文件的逻辑"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in CSV_EXTS and ext not in EXCEL_EXTS:
        print(f"错误: 不支持的文件类型 {ext}")
        return

    try:
        line_count = count_lines(file_path)
        print(f"\n📄 单文件统计:")
        print("-" * 70)
        print(f"{'文件名':<50} | {'行数':>15}")
        print("-" * 70)
        print(f"{os.path.basename(file_path):<50} | {line_count:>15,}")
        print("-" * 70)
    except Exception as e:
        print(f"读取失败: {file_path}, 错误: {e}")


def main():
    parser = argparse.ArgumentParser(description="统计 CSV / Excel 文件的行数（支持单个文件或整个目录）")
    parser.add_argument("path", help="目标目录或文件路径")
    args = parser.parse_args()
    input_path = args.path

    # 1. 检查路径是否存在
    if not os.path.exists(input_path):
        print(f"错误: {input_path} 不存在")
        return

    # 2. 判断是文件还是目录
    if os.path.isfile(input_path):
        process_single_file(input_path)
        return

    # 3. 如果是目录，执行原有的遍历逻辑
    dir_data = defaultdict(list)
    for root, _, files in os.walk(input_path):
        for name in files:
            ext = os.path.splitext(name)[1].lower()
            if ext in CSV_EXTS or ext in EXCEL_EXTS:
                file_path = os.path.join(root, name)
                try:
                    line_count = count_lines(file_path)
                    dir_data[root].append((name, line_count))
                except Exception as e:
                    dir_data[root].append((name, f"读取失败: {e}"))

    if not dir_data:
        print("未在目录中找到 CSV / Excel 文件")
        return

    overall_total = 0
    for folder in sorted(dir_data.keys()):
        print(f"\n📂 目录: {folder}")
        print("-" * 70)
        print(f"{'文件名':<50} | {'行数':>15}")
        print("-" * 70)
        
        folder_total = 0
        for file_name, count in sorted(dir_data[folder]):
            if isinstance(count, int):
                print(f"{file_name:<50} | {count:>15,}")
                folder_total += count
            else:
                print(f"{file_name:<50} | {count:>15}")
        
        print("-" * 70)
        print(f"{'目录小计':<50} | {folder_total:>15,}")
        overall_total += folder_total
        print("=" * 70)

    print(f"\n🚀 所有目录总计行数: {overall_total:,}")


if __name__ == "__main__":
    main()